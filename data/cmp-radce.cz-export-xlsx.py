#!/usr/bin/env python
# coding: utf-8

# In[15]:


import os
import openpyxl
import shutil
import os.path
from os import path


# In[16]:


DATA_INPUT_XLSX='cmp-radce.cz.xlsx'
DATA_OUTPUT_FOLDER='MODIFIED'
DATA_OUTPUT_TRAIN_FOLDER='{}/train'.format(DATA_OUTPUT_FOLDER)

COLUMN_SOURCE=0
COLUMN_FORUM_TOPIC=1
COLUMN_TEXT=2
COLUMN_CATEGORY=3


# In[17]:


if path.exists(DATA_OUTPUT_FOLDER):
    shutil.rmtree(DATA_OUTPUT_FOLDER)
os.makedirs(DATA_OUTPUT_TRAIN_FOLDER)


# In[5]:


wb_obj = openpyxl.load_workbook(DATA_INPUT_XLSX) 
sheet = wb_obj.active


# In[24]:


for i, row in enumerate(sheet.iter_rows(max_row=sheet.max_row)):
    if i==0:
        continue
    source=row[COLUMN_SOURCE].value
    forum_topic=row[COLUMN_FORUM_TOPIC].value
    text=row[COLUMN_TEXT].value
    category=row[COLUMN_CATEGORY].value
    
    category_folder='{}/{}'.format(DATA_OUTPUT_TRAIN_FOLDER, category)
    if not(path.exists(category_folder)):
        os.mkdir(category_folder)
        
    idx=len(os.listdir(category_folder))+1
    document_name='{}_{}_[{}].txt'.format(str(idx), source, forum_topic)
    document_path='{}/{}'.format(category_folder, document_name)
    
    with open(document_path, 'w') as f:
        f.write(text)
    


# In[ ]:




